import json
import os
import io
import logging
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from db import get_db
from auth import get_current_admin
from models import (
    Dragon, DragonStep, User, UserDragon, CollectionGrid, UserProgress, Family,
    ErrorLog, ServiceHeartbeat, ApiRequestLog,
    SuspiciousReport, ShopItem, StageShopItem, UserInventory,
    EpicStage, EpicStageAction, EpicActionItem, EpicMoodlet,
    Treasure, UserTreasure, DonorCache,
    PricingConfig, DragonSet, PaymentOrder,
    CharacterAxis, CharacterBalance,
    EpicSubAction, EpicSubActionItem, EpicSubActionStep, EpicSubActionOutcome,
    EpicActionOutcome,
    IntroChapter,
)
from config import API_ERROR_LOG, DONOR_SYNC_INTERVAL_HOURS
from services.dragon_service import (
    get_dragons, get_dragon, create_dragon, update_dragon, delete_dragon, sync_steps_count,
)
from services.family_service import create_family as svc_create_family, update_family as svc_update_family

router = APIRouter(prefix="/api/admin", tags=["admin"], dependencies=[Depends(get_current_admin)])


# ─── Stats ───

@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    return {
        "dragons_total": db.query(Dragon).count(),
        "dragons_active": db.query(Dragon).filter(Dragon.is_active == True).count(),
        "pins_total": db.query(Dragon).filter(Dragon.pin_code != None).count(),
        "pins_active": db.query(Dragon).filter(Dragon.pin_code != None, Dragon.is_active == True).count(),
        "pins_used": 0,
        "users_total": db.query(User).count(),
        "dragons_collected_total": db.query(UserDragon).count(),
        "suspicious_total": db.query(SuspiciousReport).count(),
    }


# ─── Dragons CRUD ───

@router.get("/dragons")
def list_dragons(db: Session = Depends(get_db)):
    return get_dragons(db)


@router.get("/dragons/export")
def export_dragons(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    dragons = db.query(Dragon).filter(Dragon.is_epic == False).order_by(Dragon.id).all()
    families = {f.id: f.name for f in db.query(Family).all()}
    treasures = {t.dragon_id: t for t in db.query(Treasure).all()}
    rarity_names = {1: "Обычный", 2: "Редкий", 3: "Легендарный"}
    phase_names = {0: "Яйцо", 1: "Легенда"}

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "Драконы"
    ws_d.append([
        "ID", "Имя", "Редкость", "Тип яйца", "Кол-во шагов", "Описание",
        "Активен", "PIN", "Семейство", "Сокровище", "Описание сокровища",
    ])
    for d in dragons:
        tr = treasures.get(d.id)
        ws_d.append([
            d.id, d.name, rarity_names.get(d.rarity, d.rarity), d.egg_type, d.steps_count,
            d.description, "да" if d.is_active else "нет", d.pin_code or "",
            families.get(d.family_id, "") if d.family_id else "",
            tr.name if tr else "", tr.description if tr else "",
        ])

    ws_s = wb.create_sheet("Шаги")
    ws_s.append([
        "Дракон ID", "Дракон", "Фаза", "Номер шага", "Магическое действие",
        "Задание", "Подсказка", "Ключевое слово", "Таймаут (ч)", "Таймаут (мин)", "Норма крестиков",
    ])
    dragon_names = {d.id: d.name for d in dragons}
    dragon_ids = list(dragon_names.keys())
    steps = (
        db.query(DragonStep)
        .filter(DragonStep.dragon_id.in_(dragon_ids))
        .order_by(DragonStep.dragon_id, DragonStep.phase, DragonStep.step_number)
        .all()
        if dragon_ids else []
    )
    for s in steps:
        ws_s.append([
            s.dragon_id, dragon_names.get(s.dragon_id, ""), phase_names.get(s.phase, s.phase),
            s.step_number, s.magic_action, s.task_description, s.hint, s.keyword,
            s.timeout_hours, s.timeout_minutes, s.crosses_norm,
        ])

    for ws in (ws_d, ws_s):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"dragons_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dragons/{dragon_id}")
def get_dragon_by_id(dragon_id: int, db: Session = Depends(get_db)):
    dragon = get_dragon(db, dragon_id)
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")
    data = {c.name: getattr(dragon, c.name) for c in dragon.__table__.columns}
    treasure = db.query(Treasure).filter(Treasure.dragon_id == dragon_id).first()
    data["treasure"] = _treasure_dict(db, treasure) if treasure else None
    return data


@router.post("/dragons")
def create_dragon_route(
    name: str = Form(...),
    rarity: int = Form(...),
    egg_type: str = Form(""),
    description: str = Form(""),
    family_id: Optional[int] = Form(None),
    image: Optional[UploadFile] = File(None),
    silhouette: Optional[UploadFile] = File(None),
    steps: Optional[str] = Form(None),
    is_epic: bool = Form(False),
    db: Session = Depends(get_db),
):
    dragon = create_dragon(db, name, rarity, egg_type, description, family_id, image, silhouette, is_epic)

    if steps:
        try:
            steps_data = json.loads(steps)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid steps JSON")
        for s in steps_data:
            th = max(0, int(s.get("timeout_hours", 0) or 0))
            tm = max(0, min(59, int(s.get("timeout_minutes", 0) or 0)))
            cn = max(1, int(s.get("crosses_norm", 1000) or 1000))
            db.add(DragonStep(
                dragon_id=dragon.id,
                step_number=s.get("step_number", 0),
                magic_action=s.get("magic_action", ""),
                task_description=s.get("task_description", ""),
                hint=s.get("hint", ""),
                keyword="вышито",
                timeout_hours=th,
                timeout_minutes=tm,
                crosses_norm=cn,
            ))
        db.commit()
        sync_steps_count(db, dragon.id)
        logging.getLogger("uvicorn").info(f"[STEPS] Saved {len(steps_data)} steps for dragon {dragon.id}")
    else:
        logging.getLogger("uvicorn").info(f"[STEPS] No steps received for dragon {dragon.id}")

    return dragon


@router.put("/dragons/{dragon_id}")
def update_dragon_route(
    dragon_id: int,
    name: Optional[str] = Form(None),
    rarity: Optional[int] = Form(None),
    egg_type: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    is_active: Optional[bool] = Form(None),
    family_id: Optional[int] = Form(None),
    image: Optional[UploadFile] = File(None),
    silhouette: Optional[UploadFile] = File(None),
    steps: Optional[str] = Form(None),
    is_epic: Optional[bool] = Form(None),
    db: Session = Depends(get_db),
):
    dragon = update_dragon(db, dragon_id, name, rarity, egg_type, description,
                           is_active, family_id, image, silhouette, is_epic)
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")

    if steps:
        try:
            steps_data = json.loads(steps)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid steps JSON")
        submitted_ids = {s.get("id", 0) for s in steps_data if s.get("id", 0) != 0}
        if submitted_ids:
            db.query(DragonStep).filter(
                DragonStep.dragon_id == dragon_id,
                DragonStep.phase == 0,
                DragonStep.id.notin_(submitted_ids),
            ).delete(synchronize_session=False)
        else:
            db.query(DragonStep).filter(
                DragonStep.dragon_id == dragon_id, DragonStep.phase == 0
            ).delete(synchronize_session=False)
        for s in steps_data:
            th = max(0, int(s.get("timeout_hours", 0) or 0))
            tm = max(0, min(59, int(s.get("timeout_minutes", 0) or 0)))
            cn = max(1, int(s.get("crosses_norm", 1000) or 1000))
            sid = s.get("id", 0)
            if sid == 0:
                step = DragonStep(dragon_id=dragon_id, step_number=s.get("step_number", 0),
                                  magic_action=s.get("magic_action", ""),
                                  task_description=s.get("task_description", ""),
                                  hint=s.get("hint", ""), keyword="вышито",
                                  timeout_hours=th, timeout_minutes=tm,
                                  crosses_norm=cn)
                db.add(step)
            else:
                step = db.query(DragonStep).filter(DragonStep.id == sid, DragonStep.dragon_id == dragon_id).first()
                if step:
                    step.step_number = s.get("step_number", step.step_number)
                    step.magic_action = s.get("magic_action", step.magic_action)
                    step.task_description = s.get("task_description", step.task_description)
                    step.hint = s.get("hint", step.hint)
                    step.keyword = "вышито"
                    step.timeout_hours = th
                    step.timeout_minutes = tm
                    step.crosses_norm = cn
        db.commit()
        sync_steps_count(db, dragon_id)

    return dragon


@router.delete("/dragons/{dragon_id}")
def delete_dragon_route(dragon_id: int, db: Session = Depends(get_db)):
    ok = delete_dragon(db, dragon_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Dragon not found")
    return {"ok": True}


# ─── Steps ───

@router.get("/dragons/{dragon_id}/steps")
def get_steps(dragon_id: int, db: Session = Depends(get_db)):
    return (
        db.query(DragonStep)
        .filter(DragonStep.dragon_id == dragon_id, DragonStep.phase == 0)
        .order_by(DragonStep.step_number)
        .all()
    )


@router.put("/dragons/{dragon_id}/steps")
async def save_steps(dragon_id: int, request: Request, db: Session = Depends(get_db)):
    raw = await request.body()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    steps_data = data.get("steps", [])
    submitted_ids = {s.get("id", 0) for s in steps_data if s.get("id", 0) != 0}
    if submitted_ids:
        db.query(DragonStep).filter(
            DragonStep.dragon_id == dragon_id,
            DragonStep.phase == 0,
            DragonStep.id.notin_(submitted_ids),
        ).delete(synchronize_session=False)
    else:
        db.query(DragonStep).filter(
            DragonStep.dragon_id == dragon_id, DragonStep.phase == 0
        ).delete(synchronize_session=False)
    for s in steps_data:
        th = max(0, int(s.get("timeout_hours", 0) or 0))
        tm = max(0, min(59, int(s.get("timeout_minutes", 0) or 0)))
        cn = max(1, int(s.get("crosses_norm", 1000) or 1000))
        sid = s.get("id", 0)
        if sid == 0:
            step = DragonStep(dragon_id=dragon_id, step_number=s.get("step_number", 0),
                              magic_action=s.get("magic_action", ""),
                              task_description=s.get("task_description", ""),
                              hint=s.get("hint", ""), keyword="вышито",
                              timeout_hours=th, timeout_minutes=tm,
                              crosses_norm=cn, image_path=s.get("image_path", ""))
            db.add(step)
        else:
            step = db.query(DragonStep).filter(DragonStep.id == sid, DragonStep.dragon_id == dragon_id).first()
            if step:
                step.step_number = s.get("step_number", step.step_number)
                step.magic_action = s.get("magic_action", step.magic_action)
                step.task_description = s.get("task_description", step.task_description)
                step.hint = s.get("hint", step.hint)
                step.keyword = "вышито"
                step.timeout_hours = th
                step.timeout_minutes = tm
                step.crosses_norm = cn
                step.image_path = s.get("image_path", step.image_path)
    db.commit()
    sync_steps_count(db, dragon_id)
    return {"ok": True, "saved": len(steps_data)}


@router.post("/dragons/{dragon_id}/steps")
def add_step(dragon_id: int, db: Session = Depends(get_db)):
    dragon = get_dragon(db, dragon_id)
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")
    max_number = db.query(DragonStep).filter(DragonStep.dragon_id == dragon_id, DragonStep.phase == 0).count()
    step = DragonStep(dragon_id=dragon_id, step_number=max_number + 1, magic_action="", task_description="", hint="", timeout_hours=1, timeout_minutes=0, crosses_norm=1000)
    db.add(step)
    db.commit()
    sync_steps_count(db, dragon_id)
    db.refresh(step)
    return step


@router.delete("/dragons/{dragon_id}/steps/{step_id}")
def delete_step(dragon_id: int, step_id: int, db: Session = Depends(get_db)):
    step = db.query(DragonStep).filter(DragonStep.id == step_id, DragonStep.dragon_id == dragon_id).first()
    if not step:
        raise HTTPException(status_code=404, detail="Step not found")
    db.delete(step)
    db.commit()
    # Re-number remaining steps
    remaining = db.query(DragonStep).filter(DragonStep.dragon_id == dragon_id, DragonStep.phase == 0).order_by(DragonStep.step_number).all()
    for i, s in enumerate(remaining):
        s.step_number = i + 1
    db.commit()
    sync_steps_count(db, dragon_id)
    return {"ok": True}


# ─── Grid ───

@router.get("/grid")
def get_grid(family_id: int = Query(...), db: Session = Depends(get_db)):
    return db.query(CollectionGrid).filter(CollectionGrid.family_id == family_id).order_by(CollectionGrid.cell_y, CollectionGrid.cell_x).all()


@router.post("/grid/create")
def create_grid(family_id: int = Query(...), columns: int = Query(...), rows: int = Query(...), db: Session = Depends(get_db)):
    db.query(CollectionGrid).filter(CollectionGrid.family_id == family_id).delete(synchronize_session=False)
    for y in range(rows):
        for x in range(columns):
            cell = CollectionGrid(family_id=family_id, cell_x=x, cell_y=y)
            db.add(cell)
    db.commit()
    return {"columns": columns, "rows": rows, "cells": columns * rows}


@router.post("/grid/resize")
def resize_grid(family_id: int = Query(...), columns: int = Query(...), rows: int = Query(...), db: Session = Depends(get_db)):
    cells = db.query(CollectionGrid).filter(CollectionGrid.family_id == family_id).order_by(CollectionGrid.cell_y, CollectionGrid.cell_x).all()
    if columns < 1 or rows < 1:
        raise HTTPException(status_code=400, detail="Min size is 1×1")

    for c in cells:
        if (c.cell_x >= columns or c.cell_y >= rows) and c.dragon_id is not None:
            raise HTTPException(status_code=400,
                detail=f"Нельзя удалить ячейку ({c.cell_x},{c.cell_y}) — в ней дракон. Уберите его сначала.")

    db.query(CollectionGrid).filter(
        CollectionGrid.family_id == family_id,
        (CollectionGrid.cell_x >= columns) | (CollectionGrid.cell_y >= rows)
    ).delete(synchronize_session=False)

    for y in range(rows):
        for x in range(columns):
            exists = db.query(CollectionGrid).filter(
                CollectionGrid.family_id == family_id,
                CollectionGrid.cell_x == x, CollectionGrid.cell_y == y
            ).first()
            if not exists:
                db.add(CollectionGrid(family_id=family_id, cell_x=x, cell_y=y))

    db.commit()
    return {"columns": columns, "rows": rows, "cells": columns * rows}


@router.put("/grid/cell/{cell_id}")
def update_cell(cell_id: int, dragon_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    cell = db.query(CollectionGrid).filter(CollectionGrid.id == cell_id).first()
    if not cell:
        raise HTTPException(status_code=404, detail="Cell not found")
    if dragon_id is not None:
        existing = db.query(CollectionGrid).filter(CollectionGrid.dragon_id == dragon_id).first()
        if existing and existing.id != cell_id:
            existing.dragon_id = None
            db.flush()
    cell.dragon_id = dragon_id
    db.commit()
    return cell


@router.delete("/grid/cell/{cell_id}")
def clear_cell(cell_id: int, db: Session = Depends(get_db)):
    cell = db.query(CollectionGrid).filter(CollectionGrid.id == cell_id).first()
    if not cell:
        raise HTTPException(status_code=404, detail="Cell not found")
    cell.dragon_id = None
    db.commit()
    return {"ok": True}



# ─── Families ───

@router.get("/families")
def list_families(db: Session = Depends(get_db)):
    families = db.query(Family).order_by(Family.sort_order, Family.id).all()
    result = []
    for f in families:
        dragon_count = db.query(Dragon).filter(Dragon.family_id == f.id).count()
        result.append({"id": f.id, "name": f.name, "description": f.description, "sort_order": f.sort_order, "color": f.color, "image_path": f.image_path or "", "dragon_count": dragon_count})
    return result


@router.post("/families")
def create_family(
    name: str = Form(...),
    description: str = Form(""),
    sort_order: int = Form(0),
    color: str = Form("#9b6fc7"),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    if not name.strip():
        raise HTTPException(status_code=400, detail="Name is required")
    return svc_create_family(db, name=name, description=description,
                             sort_order=sort_order, color=color, image=image)


@router.put("/families/{family_id}")
def update_family(
    family_id: int,
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    sort_order: Optional[int] = Form(None),
    color: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    fam = svc_update_family(db, family_id, name=name, description=description,
                            sort_order=sort_order, color=color, image=image)
    if not fam:
        raise HTTPException(status_code=404, detail="Family not found")
    return fam


@router.delete("/families/{family_id}")
def delete_family(family_id: int, db: Session = Depends(get_db)):
    fam = db.query(Family).filter(Family.id == family_id).first()
    if not fam:
        raise HTTPException(status_code=404, detail="Family not found")
    db.delete(fam)
    db.commit()
    return {"ok": True}


# ─── Pins (per-dragon summary) ───

@router.get("/pins")
def list_pins(dragon_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Dragon).filter(Dragon.pin_code != None)
    if dragon_id:
        q = q.filter(Dragon.id == dragon_id)
    return q.order_by(Dragon.id).all()


# ─── Users ───

def _resolve_vk_names(vk_ids: list[int]) -> dict[int, dict]:
    """Batch-resolve VK user names using group token. Returns {vk_id: {first_name, last_name}}."""
    import config
    if not config.VK_GROUP_TOKEN or not vk_ids:
        return {}
    try:
        import vk_api
        vk = vk_api.VkApi(token=config.VK_GROUP_TOKEN, api_version="5.199").get_api()
        ids_str = ",".join(str(x) for x in vk_ids[:1000])
        users = vk.users.get(user_ids=ids_str, fields="first_name,last_name")
        return {
            u["id"]: {
                "first_name": u.get("first_name", ""),
                "last_name": u.get("last_name", ""),
            }
            for u in users
        }
    except Exception:
        return {}


@router.get("/users")
def list_users(db: Session = Depends(get_db)):
    from sqlalchemy import func
    users = db.query(User).order_by(User.registered_at.desc()).limit(200).all()
    vk_names = _resolve_vk_names([u.vk_id for u in users])
    susp_rows = (
        db.query(SuspiciousReport.user_id, func.count(SuspiciousReport.id))
        .filter(SuspiciousReport.status == "pending")
        .group_by(SuspiciousReport.user_id)
        .all()
    )
    susp_map = {uid: cnt for uid, cnt in susp_rows}
    don_ids = {
        row[0] for row in db.query(DonorCache.vk_id).filter(DonorCache.is_don == True).all()
    }
    result = []
    for u in users:
        collected = db.query(UserDragon).filter(UserDragon.user_id == u.vk_id).count()
        nm = vk_names.get(u.vk_id, {})
        result.append({
            "vk_id": u.vk_id,
            "first_name": nm.get("first_name", ""),
            "last_name": nm.get("last_name", ""),
            "state": u.state,
            "registered_at": u.registered_at,
            "pins_activated": 0,
            "last_pin_code": None,
            "dragons_collected": collected,
            "current_dragon_id": u.current_dragon_id,
            "current_step": u.current_step,
            "suspicious_pending": susp_map.get(u.vk_id, 0),
            "is_don": u.vk_id in don_ids,
            "custom_price_per_dragon": u.custom_price_per_dragon,
        })
    return result


@router.get("/users/{vk_id}")
def get_user_detail(vk_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    pins = []

    dragons_list = []
    from services.epic_service import get_epic_name_for
    for d in db.query(Dragon).filter(Dragon.is_active == True).all():
        collected = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.dragon_id == d.id).first()
        progress = db.query(UserProgress).filter(
            UserProgress.user_id == vk_id, UserProgress.dragon_id == d.id, UserProgress.completed == True
        ).count()
        if not collected and progress == 0:
            continue
        if collected:
            if collected.completed_at:
                status = "completed"
                pct = 100
            else:
                status = "growing"
                pct = round((progress / max(d.steps_count, 1)) * 100)
        elif progress > 0:
            status = "growing"
            pct = round((progress / max(d.steps_count, 1)) * 100)
        else:
            status = "locked"
            pct = 0
        epic_custom_name = get_epic_name_for(db, vk_id, d.id) if d.is_epic else ""
        dragons_list.append({
            "dragon_id": d.id,
            "name": d.name if status == "completed" else None,
            "egg_type": d.egg_type,
            "is_epic": bool(d.is_epic),
            "epic_name": epic_custom_name,
            "status": status,
            "progress_pct": pct,
            "completed_at": collected.completed_at if collected else None,
        })

    collected_count = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.completed_at != "").count()
    active_count = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.completed_at == "").count()
    vk_nm = _resolve_vk_names([vk_id]).get(vk_id, {})

    donor = db.query(DonorCache).filter(DonorCache.vk_id == vk_id).first()

    suspicious = (
        db.query(SuspiciousReport)
        .filter(SuspiciousReport.user_id == vk_id)
        .order_by(SuspiciousReport.id.desc())
        .all()
    )

    collected_treasures = (
        db.query(Treasure)
        .join(UserTreasure, UserTreasure.treasure_id == Treasure.id)
        .filter(UserTreasure.user_id == vk_id)
        .order_by(Treasure.id)
        .all()
    )

    from services.epic_service import get_epic_name
    epic_name = get_epic_name(db, vk_id)

    return {
        "vk_id": user.vk_id,
        "first_name": vk_nm.get("first_name", ""),
        "last_name": vk_nm.get("last_name", ""),
        "registered_at": user.registered_at,
        "stitches_balance": user.stitches_balance,
        "epic_unlocked": user.epic_unlocked,
        "epic_dragon_id": user.epic_dragon_id,
        "epic_name": epic_name,
        "is_don": bool(donor.is_don) if donor else False,
        "don_since": donor.don_since if donor else None,
        "don_synced_at": donor.last_synced_at if donor else None,
        "custom_price_per_dragon": user.custom_price_per_dragon,
        "pins_activated": len(pins),
        "pins": pins,
        "dragons": dragons_list,
        "dragons_collected": collected_count,
        "dragons_active": active_count,
        "dragons_total": db.query(Dragon).filter(Dragon.is_active == True).count(),
        "suspicious_reports": [
            {
                "id": s.id,
                "dragon_id": s.dragon_id,
                "step_number": s.step_number,
                "declared_crosses": s.declared_crosses,
                "normal_crosses": s.normal_crosses,
                "mode": s.mode,
                "status": s.status,
                "created_at": s.created_at,
            }
            for s in suspicious
        ],
        "treasures_collected": [_treasure_dict(db, t) for t in collected_treasures],
    }


@router.get("/users/{vk_id}/steps")
def get_user_steps(vk_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user or not user.current_dragon_id:
        return {"dragon_name": "", "total": 0, "current_step": 0, "steps": []}

    dragon = db.query(Dragon).filter(Dragon.id == user.current_dragon_id).first()
    if not dragon:
        return {"dragon_name": "", "total": 0, "current_step": 0, "steps": []}

    steps = db.query(DragonStep).filter(DragonStep.dragon_id == dragon.id).order_by(DragonStep.step_number).all()
    result = []
    for s in steps:
        progress = db.query(UserProgress).filter(
            UserProgress.user_id == vk_id,
            UserProgress.dragon_id == dragon.id,
            UserProgress.step_number == s.step_number,
        ).first()
        result.append({
            "step_number": s.step_number,
            "task_description": s.task_description,
            "magic_action": s.magic_action,
            "hint": s.hint,
            "completed": progress.completed if progress else False,
            "current": s.step_number == user.current_step,
        })
    return {"dragon_name": dragon.name, "total": dragon.steps_count, "current_step": user.current_step, "steps": result}


@router.get("/users/{vk_id}/dragons/{dragon_id}/steps")
def get_user_dragon_steps(vk_id: int, dragon_id: int, db: Session = Depends(get_db)):
    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        return {"dragon_name": "", "total": 0, "current_step": 0, "steps": []}

    user = db.query(User).filter(User.vk_id == vk_id).first()
    current_step = 0
    if user and user.current_dragon_id == dragon_id:
        current_step = user.current_step

    steps = db.query(DragonStep).filter(DragonStep.dragon_id == dragon_id).order_by(DragonStep.step_number).all()
    result = []
    for s in steps:
        progress = db.query(UserProgress).filter(
            UserProgress.user_id == vk_id,
            UserProgress.dragon_id == dragon_id,
            UserProgress.step_number == s.step_number,
        ).first()
        result.append({
            "step_number": s.step_number,
            "task_description": s.task_description,
            "magic_action": s.magic_action,
            "hint": s.hint,
            "completed": progress.completed if progress else False,
            "current": s.step_number == current_step,
        })
    return {"dragon_name": dragon.name, "total": dragon.steps_count, "current_step": current_step, "steps": result}


@router.post("/users/{vk_id}/steps/{step_number}/toggle")
async def toggle_user_step(vk_id: int, step_number: int, request: Request, db: Session = Depends(get_db)):
    body = {}
    try:
        body = json.loads(await request.body())
    except Exception:
        pass
    dragon_id = body.get("dragon_id")
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not dragon_id:
        if not user.current_dragon_id:
            raise HTTPException(status_code=400, detail="No active dragon")
        dragon_id = user.current_dragon_id

    progress = db.query(UserProgress).filter(
        UserProgress.user_id == vk_id,
        UserProgress.dragon_id == dragon_id,
        UserProgress.step_number == step_number,
    ).first()

    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    total = dragon.steps_count

    if progress:
        was_completed = progress.completed
        progress.completed = not progress.completed
        if not progress.completed:
            progress.completed_at = ""
    else:
        was_completed = False
        progress = UserProgress(
            user_id=vk_id,
            dragon_id=dragon_id,
            step_number=step_number,
            completed=True,
            completed_at=datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        )
        db.add(progress)

    # Cascade: marking step N as completed → mark all < N as completed too
    # Cascade: marking step N as incomplete → mark all > N as incomplete too
    db.flush()
    now_ts = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    if progress.completed and not was_completed:
        for s in range(1, step_number):
            existing = db.query(UserProgress).filter(
                UserProgress.user_id == vk_id,
                UserProgress.dragon_id == dragon_id,
                UserProgress.step_number == s,
            ).first()
            if existing:
                if not existing.completed:
                    existing.completed = True
                    existing.completed_at = existing.completed_at or now_ts
            else:
                db.add(UserProgress(
                    user_id=vk_id, dragon_id=dragon_id, step_number=s,
                    completed=True, completed_at=now_ts,
                ))

    elif not progress.completed and was_completed:
        for s in range(step_number + 1, total + 1):
            existing = db.query(UserProgress).filter(
                UserProgress.user_id == vk_id,
                UserProgress.dragon_id == dragon_id,
                UserProgress.step_number == s,
            ).first()
            if existing and existing.completed:
                existing.completed = False
                existing.completed_at = ""

    db.flush()

    ud_toggle = db.query(UserDragon).filter(
        UserDragon.user_id == vk_id, UserDragon.dragon_id == dragon_id
    ).first()
    if ud_toggle:
        ud_toggle.next_step_available_at = None
        ud_toggle.timeout_notified = False

    completed_count = db.query(UserProgress).filter(
        UserProgress.user_id == vk_id,
        UserProgress.dragon_id == dragon_id,
        UserProgress.completed == True,
    ).count()

    if completed_count >= total:
        if dragon.is_epic:
            user.state = "await_epic_name"
            db.commit()
            _notify_user(
                vk_id,
                "🎉🐲 Эпическое яйцо вылупилось!\n\n"
                "🐲 Твой эпический дракон вылупился!\n"
                "Как ты его назовёшь? Напиши имя одним сообщением.",
            )
            return {"ok": True, "completed": progress.completed, "current_step": user.current_step}

        ud = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.dragon_id == dragon_id).first()
        if ud and not ud.completed_at:
            ud.completed_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        user.state = "idle"
        user.current_dragon_id = None
        user.current_step = 0

        from bot.services.grow_service import award_treasure, award_family_treasures
        from services.epic_service import maybe_spawn_first_epic
        db.commit()
        treasure = award_treasure(db, vk_id, dragon_id)
        family_treasures = award_family_treasures(db, vk_id)
        epic = maybe_spawn_first_epic(db, vk_id)

        msg = (
            f"🎉 Поздравляю! Ты вырастил дракона!\n\n"
            f"⭐ {dragon.name} ⭐\n"
            f"Редкость: {({1: 'обычный', 2: 'редкий', 3: 'легендарный'}).get(dragon.rarity, 'легендарный')} {'⭐' * min(dragon.rarity, 3)}\n"
        )
        if dragon.family_id:
            family_row = db.query(Family).filter(Family.id == dragon.family_id).first()
            if family_row:
                msg += f"Коллекция: {family_row.name}\n"
        if dragon.description:
            msg += f"\n{dragon.description}\n"
        msg += "\nЗагляни в мини-приложение, чтобы увидеть его в своей коллекции!"

        attachment = ""
        if dragon.dragon_path:
            img_path = os.path.join(os.path.dirname(__file__), "..", "..", "images", "dragons", os.path.basename(dragon.dragon_path))
            attachment = _upload_vk_image(os.path.abspath(img_path), peer_id=vk_id)

        idle_kb = json.dumps({
            "one_time": False,
            "buttons": [
                [{"action": {"type": "text", "label": "🥚 Добавить яйцо дракона", "payload": json.dumps({"cmd": "pin"}, ensure_ascii=False)}, "color": "primary"}],
                [{"action": {"type": "text", "label": "🔄🥚 Сменить яйцо дракона", "payload": json.dumps({"cmd": "garden"}, ensure_ascii=False)}, "color": "secondary"},
                 {"action": {"type": "text", "label": "❓ Помощь", "payload": json.dumps({"cmd": "help"}, ensure_ascii=False)}, "color": "secondary"}],
                [{"action": {"type": "open_link", "label": "📖 Мой Бестиарий", "link": "https://vk.com/app54663330"}}],
            ],
        }, ensure_ascii=False)

        _notify_user(vk_id, msg, attachment, keyboard=idle_kb)
        if treasure:
            t_msg = f"💎 В твоей пещере появилось новое сокровище!\nПосмотри его в мини-приложении Мой Бестиарий.\n\nПолучено: {treasure.name}"
            if treasure.description:
                t_msg += f"\n{treasure.description}"
            t_attach = ""
            if treasure.image_path:
                t_path = os.path.join(os.path.dirname(__file__), "..", "..", "images", "dragons", os.path.basename(treasure.image_path))
                t_attach = _upload_vk_image(os.path.abspath(t_path), peer_id=vk_id)
            _notify_user(vk_id, t_msg, t_attach)
        for ft in (family_treasures or []):
            ft_msg = f"💎 В твоей пещере появилось новое сокровище!\nПосмотри его в мини-приложении Мой Бестиарий.\n\nСокровище семьи: {ft.name}"
            if ft.description:
                ft_msg += f"\n{ft.description}"
            ft_attach = ""
            if ft.image_path:
                ft_path = os.path.join(os.path.dirname(__file__), "..", "..", "images", "dragons", os.path.basename(ft.image_path))
                ft_attach = _upload_vk_image(os.path.abspath(ft_path), peer_id=vk_id)
            _notify_user(vk_id, ft_msg, ft_attach)
        if epic:
            from services.epic_service import get_epic_name
            epic_name = get_epic_name(db, vk_id) or epic.egg_type or "Эпический дракон"
            e_attach = ""
            if epic.egg_path:
                e_path = os.path.join(os.path.dirname(__file__), "..", "..", "images", "dragons", os.path.basename(epic.egg_path))
                e_attach = _upload_vk_image(os.path.abspath(e_path), peer_id=vk_id)
            _notify_user(vk_id,
                f"🐲🥚 В твоей пещере появилось эпическое яйцо!\n\n"
                f"«{epic_name}» ждёт своего часа.\n"
                f"Напиши «эпический» чтобы начать уход за ним.",
                attachment=e_attach)
    else:
        user.current_step = completed_count + 1
        user.state = f"grow_step_{user.current_step}"
        step_def = db.query(DragonStep).filter(
            DragonStep.dragon_id == dragon_id,
            DragonStep.step_number == user.current_step,
        ).first()
        steps_msg = _format_step_text(step_def, user.current_step, total)
        header = f"🥚 {dragon.name}\n"
        instruction = "\nПришли 2 фото (до и после) и напиши «вышито» когда выполнишь."
        if progress.completed:
            _notify_user(vk_id, f"{header}✅ Администратор отметил шаг {step_number} как выполненный.\n\n{steps_msg}{instruction}")
        else:
            _notify_user(vk_id, f"{header}↩ Администратор отменил шаг {step_number}.\n\n{steps_msg}{instruction}")

    db.commit()
    return {"ok": True, "completed": progress.completed, "current_step": user.current_step}


def _notify_user(vk_id: int, message: str, attachment: str = "", keyboard: str = None):
    """Send a notification message to user via VK bot."""
    try:
        import config
        import random
        if not config.VK_GROUP_TOKEN:
            return
        import vk_api
        vk = vk_api.VkApi(token=config.VK_GROUP_TOKEN, api_version="5.199").get_api()
        kwargs = {
            "user_id": vk_id,
            "message": message,
            "random_id": random.randint(1, 2**31 - 1),
        }
        if attachment:
            kwargs["attachment"] = attachment
        if keyboard:
            kwargs["keyboard"] = keyboard
        vk.messages.send(**kwargs)
    except Exception:
        pass


def _upload_vk_image(filepath: str, peer_id: int = 0) -> str:
    """Upload local image to VK and return attachment string."""
    import config
    if not config.VK_GROUP_TOKEN or not filepath or not os.path.isfile(filepath):
        return ""
    import vk_api
    import requests
    vk = vk_api.VkApi(token=config.VK_GROUP_TOKEN, api_version="5.199").get_api()
    for attempt in range(3):
        try:
            upload_url = vk.photos.getMessagesUploadServer(peer_id=peer_id)["upload_url"]
            with open(filepath, "rb") as f:
                resp = requests.post(upload_url, files={"photo": ("image.jpg", f, "image/jpeg")}, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            saved = vk.photos.saveMessagesPhoto(photo=data["photo"], server=data["server"], hash=data["hash"])[0]
            return f"photo{saved['owner_id']}_{saved['id']}"
        except Exception:
            if attempt < 2:
                import time
                time.sleep(1)
    return ""


def _format_step_text(step_def, step_num: int, total: int) -> str:
    if not step_def:
        return f"📋 Шаг {step_num} из {total}"
    lines = [f"📋 Шаг {step_num} из {total}"]
    if step_def.magic_action:
        lines.append(f"✨ {step_def.magic_action}")
    if step_def.task_description:
        lines.append(f"📝 {step_def.task_description}")
    if step_def.hint:
        lines.append(f"💡 {step_def.hint}")
    return "\n".join(lines)


@router.post("/users/{vk_id}/skip-step")
async def skip_step(vk_id: int, request: Request, db: Session = Depends(get_db)):
    body = {}
    try:
        body = json.loads(await request.body())
    except Exception:
        pass
    dragon_id = body.get("dragon_id")

    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not dragon_id:
        if not user.current_dragon_id:
            raise HTTPException(status_code=400, detail="No active dragon — игрок сейчас не выращивает дракона")
        dragon_id = user.current_dragon_id

    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        raise HTTPException(status_code=400, detail="Dragon not found")

    total = dragon.steps_count

    completed = db.query(UserProgress).filter(
        UserProgress.user_id == vk_id,
        UserProgress.dragon_id == dragon_id,
        UserProgress.completed == True,
    ).count()
    next_step = completed + 1

    existing = db.query(UserProgress).filter(
        UserProgress.user_id == vk_id,
        UserProgress.dragon_id == dragon_id,
        UserProgress.step_number == next_step,
    ).first()
    if not existing:
        up = UserProgress(user_id=vk_id, dragon_id=dragon_id, step_number=next_step, completed=True)
        db.add(up)

    ud = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.dragon_id == dragon.id).first()
    if ud:
        ud.next_step_available_at = None
        ud.timeout_notified = False

    if next_step >= total:
        if dragon.is_epic:
            user.state = "await_epic_name"
            db.commit()
            _notify_user(
                vk_id,
                "🎉🐲 Эпическое яйцо вылупилось!\n\n"
                "🐲 Твой эпический дракон вылупился!\n"
                "Как ты его назовёшь? Напиши имя одним сообщением.",
            )
            return {"ok": True, "new_step": 0}

        if ud and not ud.completed_at:
            ud.completed_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        if user.current_dragon_id == dragon_id:
            user.state = "idle"
            user.current_dragon_id = None
            user.current_step = 0
        db.commit()
        from bot.services.grow_service import award_treasure, award_family_treasures
        award_treasure(db, vk_id, dragon_id)
        award_family_treasures(db, vk_id)
        from services.epic_service import maybe_spawn_first_epic
        maybe_spawn_first_epic(db, vk_id)
        return {"ok": True, "new_step": 0}

    new_step = next_step + 1
    if user.current_dragon_id == dragon_id:
        user.current_step = new_step
        user.state = f"grow_step_{new_step}"

    db.commit()
    return {"ok": True, "new_step": new_step}


@router.post("/users/{vk_id}/reset-dragon")
async def reset_dragon(vk_id: int, request: Request, db: Session = Depends(get_db)):
    body = {}
    try:
        body = json.loads(await request.body())
    except Exception:
        pass
    dragon_id = body.get("dragon_id")

    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not dragon_id:
        if not user.current_dragon_id:
            raise HTTPException(status_code=400, detail="No active dragon — игрок сейчас не выращивает дракона")
        dragon_id = user.current_dragon_id

    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        raise HTTPException(status_code=400, detail="Dragon not found")
    dragon_name = dragon.name
    ud = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.dragon_id == dragon_id).first()
    db.query(UserProgress).filter(
        UserProgress.user_id == vk_id, UserProgress.dragon_id == dragon_id
    ).delete(synchronize_session=False)
    _delete_user_treasures(db, vk_id, dragon_id)
    if ud:
        ud.next_step_available_at = None
        ud.timeout_notified = False
    if user.current_dragon_id == dragon_id:
        user.state = "idle"
        user.current_dragon_id = None
        user.current_step = 0
    db.commit()
    _notify_user(vk_id, f"🔄 Администратор сбросил прогресс дракона «{dragon_name}».")
    return {"ok": True}


@router.post("/users/{vk_id}/dragons/{dragon_id}/restart")
def restart_dragon(vk_id: int, dragon_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")

    ud = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.dragon_id == dragon_id).first()
    if not ud:
        raise HTTPException(status_code=400, detail="У игрока нет этого дракона")

    db.query(UserProgress).filter(
        UserProgress.user_id == vk_id, UserProgress.dragon_id == dragon_id
    ).delete()

    ud.next_step_available_at = None
    ud.timeout_notified = False

    if ud.completed_at:
        ud.completed_at = ""

    user.current_dragon_id = dragon_id
    user.current_step = 1
    user.state = "grow_step_1"
    user.state_data = "{}"
    db.commit()

    total = dragon.steps_count
    first_step = db.query(DragonStep).filter(DragonStep.dragon_id == dragon_id, DragonStep.step_number == 1).first()
    steps_msg = _format_step_text(first_step, 1, total)
    _notify_user(vk_id, f"🔄 Администратор возобновил выращивание дракона «{dragon.name}»!\n\n{steps_msg}\n\nПришли 2 фото (до и после) и напиши «вышито» когда выполнишь.")
    return {"ok": True}


@router.delete("/users/{vk_id}/dragons/{dragon_id}")
def delete_user_dragon(vk_id: int, dragon_id: int, db: Session = Depends(get_db)):
    ud = db.query(UserDragon).filter(UserDragon.user_id == vk_id, UserDragon.dragon_id == dragon_id).first()
    if not ud:
        raise HTTPException(status_code=404, detail="Dragon not found for this user")

    db.query(UserProgress).filter(
        UserProgress.user_id == vk_id, UserProgress.dragon_id == dragon_id
    ).delete(synchronize_session=False)
    _delete_user_treasures(db, vk_id, dragon_id)

    user = db.query(User).filter(User.vk_id == vk_id).first()
    if user and user.current_dragon_id == dragon_id:
        user.current_dragon_id = None
        user.current_step = 0
        user.state = "idle"

    db.delete(ud)
    db.commit()
    return {"ok": True}


@router.delete("/users/{vk_id}")
def delete_user(vk_id: int, db: Session = Depends(get_db)):
    from models import (
        UserProgress, UserLegendProgress, UserDragon,
        UserTreasure, UserInventory, SuspiciousReport,
        EpicCareState, EpicMoodlet,
    )
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    ud_ids = [r[0] for r in db.query(UserDragon.id).filter(UserDragon.user_id == vk_id).all()]
    if ud_ids:
        db.query(EpicCareState).filter(EpicCareState.user_dragon_id.in_(ud_ids)).delete(synchronize_session=False)
        db.query(EpicMoodlet).filter(EpicMoodlet.user_dragon_id.in_(ud_ids)).delete(synchronize_session=False)

    db.query(UserProgress).filter(UserProgress.user_id == vk_id).delete(synchronize_session=False)
    db.query(UserLegendProgress).filter(UserLegendProgress.user_id == vk_id).delete(synchronize_session=False)
    db.query(UserTreasure).filter(UserTreasure.user_id == vk_id).delete(synchronize_session=False)
    db.query(UserInventory).filter(UserInventory.user_id == vk_id).delete(synchronize_session=False)
    db.query(SuspiciousReport).filter(SuspiciousReport.user_id == vk_id).delete(synchronize_session=False)
    db.query(UserDragon).filter(UserDragon.user_id == vk_id).delete(synchronize_session=False)

    db.delete(user)
    db.commit()
    return {"ok": True}


@router.get("/logs")
def list_logs(page: int = Query(1, ge=1), per_page: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)):
    offset = (page - 1) * per_page
    total = db.query(ErrorLog).count()
    logs = db.query(ErrorLog).order_by(ErrorLog.id.desc()).offset(offset).limit(per_page).all()
    return {"logs": logs, "total": total, "page": page, "per_page": per_page}


@router.get("/logs/api")
def list_api_logs(page: int = Query(1, ge=1), per_page: int = Query(50, ge=1, le=200)):
    if not os.path.isfile(API_ERROR_LOG):
        return {"lines": [], "total": 0, "page": page, "per_page": per_page}
    try:
        with open(API_ERROR_LOG, "r", encoding="utf-8", errors="replace") as f:
            all_lines = f.readlines()
    except Exception:
        return {"lines": [], "total": 0, "page": page, "per_page": per_page}
    all_lines.reverse()
    total = len(all_lines)
    offset = (page - 1) * per_page
    chunk = all_lines[offset:offset + per_page]
    return {"lines": chunk, "total": total, "page": page, "per_page": per_page}


@router.get("/logs/api-requests")
def list_api_requests(page: int = Query(1, ge=1), per_page: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)):
    offset = (page - 1) * per_page
    total = db.query(ApiRequestLog).count()
    items = db.query(ApiRequestLog).order_by(ApiRequestLog.id.desc()).offset(offset).limit(per_page).all()
    return {"items": items, "total": total, "page": page, "per_page": per_page}


@router.post("/logs/clear")
def clear_logs(db: Session = Depends(get_db)):
    db.query(ErrorLog).delete(synchronize_session=False)
    db.query(ApiRequestLog).delete(synchronize_session=False)
    db.commit()
    return {"ok": True}


@router.get("/donors")
def list_donors(db: Session = Depends(get_db)):
    donors = db.query(DonorCache).order_by(DonorCache.updated_at.desc()).all()
    return {"donors": donors, "total": len(donors)}


@router.get("/health")
def get_health(db: Session = Depends(get_db)):
    now = datetime.now()
    services = {}

    hb = db.query(ServiceHeartbeat).filter(ServiceHeartbeat.service_name == "bot").first()
    services["bot"] = _service_status(hb.last_seen if hb else None, now, 90)

    last_sync = db.query(func.max(DonorCache.last_synced_at)).scalar()
    donor_threshold = max(1, DONOR_SYNC_INTERVAL_HOURS) * 2 * 3600
    services["donor_sync"] = _service_status(last_sync, now, donor_threshold)

    return {"services": services, "checked_at": now.isoformat()}


def _service_status(last_seen, now: datetime, threshold_seconds: int) -> dict:
    if not last_seen:
        return {"status": "unknown", "last_seen": None}
    try:
        last = datetime.fromisoformat(last_seen)
    except ValueError:
        return {"status": "unknown", "last_seen": last_seen}
    online = (now - last).total_seconds() < threshold_seconds
    return {"status": "online" if online else "offline", "last_seen": last_seen}


# ═══════════════════════════════════════════════════════════
#  Phase 0 — новые сущности (CRUD без игровой логики)
# ═══════════════════════════════════════════════════════════

async def _json_body(request: Request) -> dict:
    try:
        return json.loads(await request.body())
    except Exception:
        return {}


def _delete_user_treasures(db: Session, vk_id: int, dragon_id: int):
    treasure_ids = [
        t.id for t in db.query(Treasure.id).filter(Treasure.dragon_id == dragon_id).all()
    ]
    if treasure_ids:
        db.query(UserTreasure).filter(
            UserTreasure.user_id == vk_id,
            UserTreasure.treasure_id.in_(treasure_ids),
        ).delete(synchronize_session=False)


def _upload_image_file(file: UploadFile, prefix: str) -> str:
    from services.dragon_service import _save_upload, IMAGES_DIR
    filename = _save_upload(file, IMAGES_DIR, prefix)
    return f"dragons/{filename}"


@router.post("/upload-image")
def upload_image(image: UploadFile = File(...)):
    if not image.filename:
        raise HTTPException(status_code=400, detail="No file")
    return {"path": _upload_image_file(image, "asset")}


# ─── Treasures (Фаза 8) ───

def _treasure_dict(db, t: Treasure) -> dict:
    dragon_name = None
    family_name = None
    if t.dragon_id:
        dragon = db.query(Dragon).filter(Dragon.id == t.dragon_id).first()
        dragon_name = dragon.name if dragon else None
    if t.family_id:
        family = db.query(Family).filter(Family.id == t.family_id).first()
        family_name = family.name if family else None
    return {
        "id": t.id,
        "name": t.name,
        "description": t.description,
        "image_path": f"/api/static/images/{t.image_path}" if t.image_path else "",
        "dragon_id": t.dragon_id,
        "dragon_name": dragon_name,
        "family_id": t.family_id,
        "family_name": family_name,
        "is_active": t.is_active,
    }


@router.get("/treasures/available-dragons")
def list_dragons_without_treasure(db: Session = Depends(get_db)):
    from sqlalchemy import select
    used_ids = select(Treasure.dragon_id).where(Treasure.is_active == True)
    dragons = (
        db.query(Dragon)
        .filter(Dragon.rarity == 2, Dragon.is_active == True, ~Dragon.id.in_(used_ids))
        .order_by(Dragon.name)
        .all()
    )
    return [{"id": d.id, "name": d.name, "egg_type": d.egg_type} for d in dragons]


@router.get("/treasures")
def list_treasures(db: Session = Depends(get_db)):
    treasures = db.query(Treasure).order_by(Treasure.id).all()
    return [_treasure_dict(db, t) for t in treasures]


@router.post("/dragons/{dragon_id}/treasure")
def upsert_dragon_treasure(
    dragon_id: int,
    name: str = Form(...),
    description: str = Form(""),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")
    if dragon.rarity != 2:
        raise HTTPException(status_code=400, detail="Сокровище доступно только у редкого дракона (rarity=2)")

    from services.dragon_service import _save_upload, _cleanup_old, IMAGES_DIR
    treasure = db.query(Treasure).filter(Treasure.dragon_id == dragon_id).first()
    if not treasure:
        treasure = Treasure(dragon_id=dragon_id, name=name, description=description)
        db.add(treasure)
    else:
        treasure.name = name
        treasure.description = description

    if image and image.filename:
        _cleanup_old(IMAGES_DIR, f"treasure_{dragon_id}", treasure.image_path)
        filename = _save_upload(image, IMAGES_DIR, f"treasure_{dragon_id}")
        treasure.image_path = f"dragons/{filename}"

    db.commit()
    db.refresh(treasure)
    return _treasure_dict(db, treasure)


@router.put("/treasures/{treasure_id}")
def update_treasure(
    treasure_id: int,
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    is_active: Optional[bool] = Form(None),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    treasure = db.query(Treasure).filter(Treasure.id == treasure_id).first()
    if not treasure:
        raise HTTPException(status_code=404, detail="Treasure not found")

    if name is not None:
        treasure.name = name
    if description is not None:
        treasure.description = description
    if is_active is not None:
        treasure.is_active = is_active

    if image and image.filename:
        from services.dragon_service import _save_upload, _cleanup_old, IMAGES_DIR
        _cleanup_old(IMAGES_DIR, f"treasure_{treasure.dragon_id}", treasure.image_path)
        filename = _save_upload(image, IMAGES_DIR, f"treasure_{treasure.dragon_id}")
        treasure.image_path = f"dragons/{filename}"

    db.commit()
    db.refresh(treasure)
    return _treasure_dict(db, treasure)


@router.delete("/treasures/{treasure_id}")
def delete_treasure(treasure_id: int, db: Session = Depends(get_db)):
    treasure = db.query(Treasure).filter(Treasure.id == treasure_id).first()
    if not treasure:
        raise HTTPException(status_code=404, detail="Treasure not found")
    db.query(UserTreasure).filter(UserTreasure.treasure_id == treasure_id).delete(synchronize_session=False)
    db.delete(treasure)
    db.commit()
    return {"ok": True}


@router.get("/treasures/available-families")
def list_families_without_treasure(db: Session = Depends(get_db)):
    used_ids = db.query(Treasure.family_id).filter(Treasure.family_id.isnot(None), Treasure.is_active == True).all()
    used = {r[0] for r in used_ids}
    families = (
        db.query(Family)
        .filter(~Family.id.in_(used) if used else True)
        .order_by(Family.sort_order, Family.id)
        .all()
    )
    return [{"id": f.id, "name": f.name, "color": f.color} for f in families]


@router.post("/families/{family_id}/treasure")
def upsert_family_treasure(
    family_id: int,
    name: str = Form(...),
    description: str = Form(""),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    family = db.query(Family).filter(Family.id == family_id).first()
    if not family:
        raise HTTPException(status_code=404, detail="Family not found")

    from services.dragon_service import _save_upload, _cleanup_old, IMAGES_DIR
    treasure = db.query(Treasure).filter(Treasure.family_id == family_id).first()
    if not treasure:
        treasure = Treasure(family_id=family_id, name=name, description=description)
        db.add(treasure)
    else:
        treasure.name = name
        treasure.description = description

    if image and image.filename:
        _cleanup_old(IMAGES_DIR, f"treasure_family_{family_id}", treasure.image_path)
        filename = _save_upload(image, IMAGES_DIR, f"treasure_family_{family_id}")
        treasure.image_path = f"dragons/{filename}"

    db.commit()
    db.refresh(treasure)
    return _treasure_dict(db, treasure)


# ─── Shop items ───

@router.get("/shop-items")
def list_shop_items(db: Session = Depends(get_db)):
    return db.query(ShopItem).order_by(ShopItem.sort_order, ShopItem.id).all()


@router.post("/shop-items")
async def create_shop_item(request: Request, db: Session = Depends(get_db)):
    b = await _json_body(request)
    name = (b.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    item = ShopItem(
        name=name,
        description=b.get("description", ""),
        cost_stitches=max(0, int(b.get("cost_stitches", 0) or 0)),
        category=b.get("category", ""),
        image_path=b.get("image_path", ""),
        is_consumable=bool(b.get("is_consumable", True)),
        sort_order=int(b.get("sort_order", 0) or 0),
        is_active=bool(b.get("is_active", True)),
        is_legend_book=bool(b.get("is_legend_book", False)),
        is_optional=bool(b.get("is_optional", False)),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/shop-items/{item_id}")
async def update_shop_item(item_id: int, request: Request, db: Session = Depends(get_db)):
    item = db.query(ShopItem).filter(ShopItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    b = await _json_body(request)
    if "name" in b: item.name = b["name"]
    if "description" in b: item.description = b["description"]
    if "cost_stitches" in b: item.cost_stitches = max(0, int(b["cost_stitches"] or 0))
    if "category" in b: item.category = b["category"]
    if "image_path" in b: item.image_path = b["image_path"]
    if "is_consumable" in b: item.is_consumable = bool(b["is_consumable"])
    if "sort_order" in b: item.sort_order = int(b["sort_order"] or 0)
    if "is_active" in b: item.is_active = bool(b["is_active"])
    if "is_legend_book" in b: item.is_legend_book = bool(b["is_legend_book"])
    if "is_optional" in b: item.is_optional = bool(b["is_optional"])
    db.commit()
    db.refresh(item)
    return item


@router.delete("/shop-items/{item_id}")
def delete_shop_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(ShopItem).filter(ShopItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


# ─── Stage ↔ shop item binding ───

@router.get("/stage-shop-items")
def list_stage_shop_items(stage_key: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(StageShopItem)
    if stage_key:
        q = q.filter(StageShopItem.stage_key == stage_key)
    return q.order_by(StageShopItem.sort_order, StageShopItem.id).all()


@router.post("/stage-shop-items")
async def create_stage_shop_item(request: Request, db: Session = Depends(get_db)):
    b = await _json_body(request)
    stage_key = (b.get("stage_key") or "").strip()
    item_id = b.get("item_id")
    if not stage_key or not item_id:
        raise HTTPException(status_code=400, detail="stage_key and item_id required")
    if not db.query(ShopItem).filter(ShopItem.id == item_id).first():
        raise HTTPException(status_code=404, detail="Item not found")
    exists = db.query(StageShopItem).filter(
        StageShopItem.stage_key == stage_key, StageShopItem.item_id == item_id
    ).first()
    if exists:
        raise HTTPException(status_code=400, detail="Already bound to this stage")
    link = StageShopItem(stage_key=stage_key, item_id=item_id, sort_order=int(b.get("sort_order", 0) or 0))
    db.add(link)
    db.commit()
    db.refresh(link)
    return link


@router.delete("/stage-shop-items/{link_id}")
def delete_stage_shop_item(link_id: int, db: Session = Depends(get_db)):
    link = db.query(StageShopItem).filter(StageShopItem.id == link_id).first()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    db.delete(link)
    db.commit()
    return {"ok": True}


# ─── Epic species (dragons with is_epic=True) ───

@router.get("/epic/species")
def list_epic_species(db: Session = Depends(get_db)):
    return db.query(Dragon).filter(Dragon.is_epic == True).order_by(Dragon.id).all()


# ─── Epic stages (общий конвейер) ───

@router.get("/epic/stages")
def list_epic_stages(db: Session = Depends(get_db)):
    return db.query(EpicStage).order_by(EpicStage.stage_number, EpicStage.id).all()


@router.post("/epic/stages")
async def create_epic_stage(request: Request, db: Session = Depends(get_db)):
    b = await _json_body(request)
    stage = EpicStage(
        stage_number=int(b.get("stage_number", 0) or 0),
        name=b.get("name", ""),
        description=b.get("description", ""),
        image_start=b.get("image_start", ""),
        image_end=b.get("image_end", ""),
        cycles_count=max(1, int(b.get("cycles_count", 3) or 3)),
    )
    db.add(stage)
    db.commit()
    db.refresh(stage)
    return stage


@router.put("/epic/stages/{stage_id}")
async def update_epic_stage(stage_id: int, request: Request, db: Session = Depends(get_db)):
    stage = db.query(EpicStage).filter(EpicStage.id == stage_id).first()
    if not stage:
        raise HTTPException(status_code=404, detail="Stage not found")
    b = await _json_body(request)
    if "stage_number" in b: stage.stage_number = int(b["stage_number"] or 0)
    if "name" in b: stage.name = b["name"]
    if "description" in b: stage.description = b["description"]
    if "image_start" in b: stage.image_start = b["image_start"]
    if "image_end" in b: stage.image_end = b["image_end"]
    if "cycles_count" in b: stage.cycles_count = max(1, int(b["cycles_count"] or 3))
    db.commit()
    db.refresh(stage)
    return stage


@router.delete("/epic/stages/{stage_id}")
def delete_epic_stage(stage_id: int, db: Session = Depends(get_db)):
    stage = db.query(EpicStage).filter(EpicStage.id == stage_id).first()
    if not stage:
        raise HTTPException(status_code=404, detail="Stage not found")
    db.delete(stage)
    db.commit()
    return {"ok": True}


# ─── Character Axes ───

@router.get("/character-axes")
def list_character_axes(db: Session = Depends(get_db)):
    return db.query(CharacterAxis).order_by(CharacterAxis.sort_order, CharacterAxis.id).all()


@router.post("/character-axes")
async def create_character_axis(request: Request, db: Session = Depends(get_db)):
    b = await _json_body(request)
    max_order = db.query(func.max(CharacterAxis.sort_order)).scalar() or 0
    axis = CharacterAxis(
        positive_label=b.get("positive_label", ""),
        negative_label=b.get("negative_label", ""),
        sort_order=max_order + 1,
        is_active=bool(b.get("is_active", True)),
    )
    db.add(axis)
    db.commit()
    db.refresh(axis)
    return axis


@router.put("/character-axes/{axis_id}")
async def update_character_axis(axis_id: int, request: Request, db: Session = Depends(get_db)):
    axis = db.query(CharacterAxis).filter(CharacterAxis.id == axis_id).first()
    if not axis:
        raise HTTPException(status_code=404, detail="Axis not found")
    b = await _json_body(request)
    if "positive_label" in b: axis.positive_label = b["positive_label"]
    if "negative_label" in b: axis.negative_label = b["negative_label"]
    if "sort_order" in b: axis.sort_order = int(b["sort_order"] or 0)
    if "is_active" in b: axis.is_active = bool(b["is_active"])
    db.commit()
    db.refresh(axis)
    return axis


@router.delete("/character-axes/{axis_id}")
def delete_character_axis(axis_id: int, db: Session = Depends(get_db)):
    axis = db.query(CharacterAxis).filter(CharacterAxis.id == axis_id).first()
    if not axis:
        raise HTTPException(status_code=404, detail="Axis not found")
    db.delete(axis)
    db.commit()
    return {"ok": True}


# ─── Epic stage actions (уход) ───

def _resync_stage_items(db, stage_id: int):
    """Синхронизирует привязку товаров к стадии по товарам действий ухода."""
    stage = db.query(EpicStage).filter(EpicStage.id == stage_id).first()
    if not stage:
        return
    key = f"epic:{stage.stage_number}"
    action_ids = [a.id for a in db.query(EpicStageAction).filter(EpicStageAction.stage_id == stage_id).all()]
    want = set()
    if action_ids:
        want = {
            ai.item_id
            for ai in db.query(EpicActionItem).filter(EpicActionItem.action_id.in_(action_ids)).all()
        }
        sub_ids = [sa.id for sa in db.query(EpicSubAction).filter(EpicSubAction.action_id.in_(action_ids)).all()]
        if sub_ids:
            want |= {
                sai.item_id
                for sai in db.query(EpicSubActionItem).filter(EpicSubActionItem.sub_action_id.in_(sub_ids)).all()
            }
    existing = db.query(StageShopItem).filter(StageShopItem.stage_key == key).all()
    have = {l.item_id: l for l in existing}
    for item_id, link in have.items():
        if item_id not in want:
            db.delete(link)
    for item_id in want:
        if item_id not in have:
            db.add(StageShopItem(stage_key=key, item_id=item_id))
    db.commit()


def _action_items(db, action_id: int) -> list[int]:
    return [ai.item_id for ai in db.query(EpicActionItem).filter(EpicActionItem.action_id == action_id).all()]


def _set_action_items(db, action_id: int, item_ids):
    db.query(EpicActionItem).filter(EpicActionItem.action_id == action_id).delete(synchronize_session=False)
    seen = set()
    for iid in (item_ids or []):
        if iid and iid not in seen and db.query(ShopItem).filter(ShopItem.id == iid).first():
            db.add(EpicActionItem(action_id=action_id, item_id=iid))
            seen.add(iid)
    db.commit()


def _action_dict(db, action: EpicStageAction) -> dict:
    result = {
        "id": action.id,
        "dragon_id": action.dragon_id,
        "stage_id": action.stage_id,
        "action_label": action.action_label,
        "order_in_cycle": action.order_in_cycle,
        "task": action.task,
        "hint": action.hint,
        "crosses_norm": action.crosses_norm,
        "image_path": action.image_path or "",
        "action_type": getattr(action, "action_type", "simple") or "simple",
        "timeout_hours": action.timeout_hours,
        "timeout_minutes": action.timeout_minutes,
        "item_ids": _action_items(db, action.id),
    }
    atype = result["action_type"]
    if atype == "composite":
        result["item_ids"] = []
        result["sub_actions"] = _sub_actions_list(db, action.id)
    else:
        result["random_outcome"] = bool(getattr(action, "random_outcome", True))
        result["character_axis_id"] = action.character_axis_id
        result["description"] = getattr(action, "description", "") or ""
        result["confirm_button_label"] = getattr(action, "confirm_button_label", "") or ""
        outcomes = db.query(EpicActionOutcome).filter(EpicActionOutcome.action_id == action.id).all()
        result["outcomes"] = [{
            "id": o.id,
            "polarity": o.polarity,
            "label": o.label,
            "moodlet_title": o.moodlet_title,
            "moodlet_text": o.moodlet_text,
            "image_path": o.image_path or "",
        } for o in outcomes]
    return result


def _sub_actions_list(db, action_id: int) -> list:
    sub_actions = db.query(EpicSubAction).filter(EpicSubAction.action_id == action_id).order_by(EpicSubAction.order_in_sub, EpicSubAction.id).all()
    result = []
    for sa in sub_actions:
        steps = db.query(EpicSubActionStep).filter(EpicSubActionStep.sub_action_id == sa.id).order_by(EpicSubActionStep.order, EpicSubActionStep.id).all()
        outcomes = db.query(EpicSubActionOutcome).filter(EpicSubActionOutcome.sub_action_id == sa.id).all()
        sai_ids = [sai.item_id for sai in db.query(EpicSubActionItem).filter(EpicSubActionItem.sub_action_id == sa.id).all()]
        result.append({
            "id": sa.id,
            "label": sa.label,
            "description": sa.description,
            "confirm_button_label": sa.confirm_button_label or "",
            "random_outcome": bool(sa.random_outcome),
            "order_in_sub": sa.order_in_sub,
            "image_path": sa.image_path or "",
            "character_axis_id": sa.character_axis_id,
            "item_ids": sai_ids,
            "steps": [{
                "id": s.id,
                "step_label": s.step_label,
                "order": s.order,
                "task": s.task,
                "hint": s.hint,
                "crosses_norm": s.crosses_norm,
                "image_path": s.image_path or "",
                "timeout_hours": s.timeout_hours,
                "timeout_minutes": s.timeout_minutes,
            } for s in steps],
            "outcomes": [{
                "id": o.id,
                "polarity": o.polarity,
                "label": o.label,
                "moodlet_title": o.moodlet_title,
                "moodlet_text": o.moodlet_text,
                "image_path": o.image_path or "",
            } for o in outcomes],
        })
    return result


@router.get("/epic/species/{dragon_id}/stages/{stage_id}/actions")
def list_epic_actions(dragon_id: int, stage_id: int, db: Session = Depends(get_db)):
    actions = (
        db.query(EpicStageAction)
        .filter(EpicStageAction.stage_id == stage_id, EpicStageAction.dragon_id == dragon_id)
        .order_by(EpicStageAction.order_in_cycle, EpicStageAction.id)
        .all()
    )
    return [_action_dict(db, a) for a in actions]


@router.post("/epic/species/{dragon_id}/stages/{stage_id}/actions")
async def create_epic_action(dragon_id: int, stage_id: int, request: Request, db: Session = Depends(get_db)):
    if not db.query(EpicStage).filter(EpicStage.id == stage_id).first():
        raise HTTPException(status_code=404, detail="Stage not found")
    if not db.query(Dragon).filter(Dragon.id == dragon_id, Dragon.is_epic == True).first():
        raise HTTPException(status_code=404, detail="Epic dragon not found")
    b = await _json_body(request)
    label = b.get("action_label", "")
    atype = b.get("action_type", "simple") or "simple"
    if atype == "composite" and b.get("item_ids"):
        raise HTTPException(status_code=422, detail="composite actions must not have direct item_ids (items belong to sub_actions)")
    action = EpicStageAction(
        dragon_id=dragon_id,
        stage_id=stage_id,
        action_label=label,
        order_in_cycle=int(b.get("order_in_cycle", 0) or 0),
        task=b.get("task", ""),
        hint=b.get("hint", ""),
        crosses_norm=max(1, int(b.get("crosses_norm", 1000) or 1000)),
        image_path=b.get("image_path", ""),
        action_type=atype,
        timeout_hours=max(0, int(b.get("timeout_hours", 0) or 0)),
        timeout_minutes=max(0, min(59, int(b.get("timeout_minutes", 0) or 0))),
        random_outcome=bool(b.get("random_outcome", True)),
        character_axis_id=int(b["character_axis_id"]) if b.get("character_axis_id") else None,
        description=b.get("description", ""),
        confirm_button_label=b.get("confirm_button_label", ""),
    )
    db.add(action)
    db.commit()
    db.refresh(action)
    if atype != "composite":
        _set_action_items(db, action.id, b.get("item_ids", []))
        for polarity in ("positive", "negative"):
            db.add(EpicActionOutcome(action_id=action.id, polarity=polarity))
        db.commit()
    _resync_stage_items(db, stage_id)
    db.refresh(action)
    return _action_dict(db, action)


@router.put("/epic/actions/{action_id}")
async def update_epic_action(action_id: int, request: Request, db: Session = Depends(get_db)):
    action = db.query(EpicStageAction).filter(EpicStageAction.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    b = await _json_body(request)
    if "action_label" in b: action.action_label = b["action_label"]
    if "order_in_cycle" in b: action.order_in_cycle = int(b["order_in_cycle"] or 0)
    if "task" in b: action.task = b["task"]
    if "hint" in b: action.hint = b["hint"]
    if "crosses_norm" in b: action.crosses_norm = max(1, int(b["crosses_norm"] or 1000))
    if "image_path" in b: action.image_path = b["image_path"]
    if "action_type" in b: action.action_type = b["action_type"]
    if "timeout_hours" in b: action.timeout_hours = max(0, int(b["timeout_hours"] or 0))
    if "timeout_minutes" in b: action.timeout_minutes = max(0, min(59, int(b["timeout_minutes"] or 0)))
    if "random_outcome" in b: action.random_outcome = bool(b["random_outcome"])
    if "character_axis_id" in b: action.character_axis_id = int(b["character_axis_id"]) if b["character_axis_id"] else None
    if "description" in b: action.description = b["description"]
    if "confirm_button_label" in b: action.confirm_button_label = b["confirm_button_label"]
    db.commit()
    db.refresh(action)
    atype = getattr(action, "action_type", "simple") or "simple"
    if atype != "composite" and "item_ids" in b:
        _set_action_items(db, action.id, b.get("item_ids", []))
    if atype == "composite" and "item_ids" in b and b.get("item_ids"):
        raise HTTPException(status_code=422, detail="composite actions must not have direct item_ids")
    if atype != "composite":
        existing = {o.polarity for o in db.query(EpicActionOutcome).filter(EpicActionOutcome.action_id == action.id).all()}
        for polarity in ("positive", "negative"):
            if polarity not in existing:
                db.add(EpicActionOutcome(action_id=action.id, polarity=polarity))
        db.commit()
    stage_id = action.stage_id
    _resync_stage_items(db, stage_id)
    db.refresh(action)
    return _action_dict(db, action)


@router.delete("/epic/actions/{action_id}")
def delete_epic_action(action_id: int, db: Session = Depends(get_db)):
    action = db.query(EpicStageAction).filter(EpicStageAction.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    stage_id = action.stage_id
    db.delete(action)
    db.commit()
    _resync_stage_items(db, stage_id)
    return {"ok": True}


# ─── Epic Sub-Actions (composite) ───

@router.get("/epic/actions/{action_id}/sub-actions")
def list_sub_actions(action_id: int, db: Session = Depends(get_db)):
    return _sub_actions_list(db, action_id)


@router.post("/epic/actions/{action_id}/sub-actions")
async def create_sub_action(action_id: int, request: Request, db: Session = Depends(get_db)):
    action = db.query(EpicStageAction).filter(EpicStageAction.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    b = await _json_body(request)
    axis_id = b.get("character_axis_id")
    sa = EpicSubAction(
        action_id=action_id,
        label=b.get("label", ""),
        description=b.get("description", ""),
        confirm_button_label=b.get("confirm_button_label", ""),
        random_outcome=bool(b.get("random_outcome", True)),
        order_in_sub=int(b.get("order_in_sub", 0) or 0),
        image_path=b.get("image_path", ""),
        character_axis_id=int(axis_id) if axis_id else None,
    )
    db.add(sa)
    db.commit()
    db.refresh(sa)
    for polarity in ("positive", "negative"):
        db.add(EpicSubActionOutcome(sub_action_id=sa.id, polarity=polarity))
    db.commit()
    _set_sub_action_items(db, sa.id, b.get("item_ids", []))
    _resync_stage_items(db, action.stage_id)
    db.refresh(sa)
    return sa


@router.put("/epic/sub-actions/{sub_id}")
async def update_sub_action(sub_id: int, request: Request, db: Session = Depends(get_db)):
    sa = db.query(EpicSubAction).filter(EpicSubAction.id == sub_id).first()
    if not sa:
        raise HTTPException(status_code=404, detail="Sub-action not found")
    b = await _json_body(request)
    if "label" in b: sa.label = b["label"]
    if "description" in b: sa.description = b["description"]
    if "confirm_button_label" in b: sa.confirm_button_label = b["confirm_button_label"]
    if "random_outcome" in b: sa.random_outcome = bool(b["random_outcome"])
    if "order_in_sub" in b: sa.order_in_sub = int(b["order_in_sub"] or 0)
    if "image_path" in b: sa.image_path = b["image_path"]
    if "character_axis_id" in b: sa.character_axis_id = int(b["character_axis_id"] or 0) or None
    db.commit()
    if "item_ids" in b:
        _set_sub_action_items(db, sa.id, b.get("item_ids", []))
        action = db.query(EpicStageAction).filter(EpicStageAction.id == sa.action_id).first()
        if action:
            _resync_stage_items(db, action.stage_id)
    db.refresh(sa)
    return sa


@router.delete("/epic/sub-actions/{sub_id}")
def delete_sub_action(sub_id: int, db: Session = Depends(get_db)):
    sa = db.query(EpicSubAction).filter(EpicSubAction.id == sub_id).first()
    if not sa:
        raise HTTPException(status_code=404, detail="Sub-action not found")
    action_id = sa.action_id
    action = db.query(EpicStageAction).filter(EpicStageAction.id == action_id).first()
    db.delete(sa)
    db.commit()
    if action:
        _resync_stage_items(db, action.stage_id)
    return {"ok": True}


def _set_sub_action_items(db, sub_action_id: int, item_ids):
    db.query(EpicSubActionItem).filter(EpicSubActionItem.sub_action_id == sub_action_id).delete(synchronize_session=False)
    seen = set()
    for iid in (item_ids or []):
        if iid and iid not in seen and db.query(ShopItem).filter(ShopItem.id == iid).first():
            db.add(EpicSubActionItem(sub_action_id=sub_action_id, item_id=iid))
            seen.add(iid)
    db.commit()


# ─── Epic Sub-Action Steps ───

@router.get("/epic/sub-actions/{sub_id}/steps")
def list_sub_steps(sub_id: int, db: Session = Depends(get_db)):
    return (
        db.query(EpicSubActionStep)
        .filter(EpicSubActionStep.sub_action_id == sub_id)
        .order_by(EpicSubActionStep.order, EpicSubActionStep.id)
        .all()
    )


@router.post("/epic/sub-actions/{sub_id}/steps")
async def create_sub_step(sub_id: int, request: Request, db: Session = Depends(get_db)):
    if not db.query(EpicSubAction).filter(EpicSubAction.id == sub_id).first():
        raise HTTPException(status_code=404, detail="Sub-action not found")
    b = await _json_body(request)
    step = EpicSubActionStep(
        sub_action_id=sub_id,
        step_label=b.get("step_label", ""),
        order=int(b.get("order", 0) or 0),
        task=b.get("task", ""),
        hint=b.get("hint", ""),
        crosses_norm=max(1, int(b.get("crosses_norm", 1000) or 1000)),
        image_path=b.get("image_path", ""),
        timeout_hours=max(0, int(b.get("timeout_hours", 0) or 0)),
        timeout_minutes=max(0, min(59, int(b.get("timeout_minutes", 0) or 0))),
    )
    db.add(step)
    db.commit()
    db.refresh(step)
    return step


@router.put("/epic/sub-actions/steps/{step_id}")
async def update_sub_step(step_id: int, request: Request, db: Session = Depends(get_db)):
    step = db.query(EpicSubActionStep).filter(EpicSubActionStep.id == step_id).first()
    if not step:
        raise HTTPException(status_code=404, detail="Step not found")
    b = await _json_body(request)
    if "step_label" in b: step.step_label = b["step_label"]
    if "order" in b: step.order = int(b["order"] or 0)
    if "task" in b: step.task = b["task"]
    if "hint" in b: step.hint = b["hint"]
    if "crosses_norm" in b: step.crosses_norm = max(1, int(b["crosses_norm"] or 1000))
    if "image_path" in b: step.image_path = b["image_path"]
    if "timeout_hours" in b: step.timeout_hours = max(0, int(b["timeout_hours"] or 0))
    if "timeout_minutes" in b: step.timeout_minutes = max(0, min(59, int(b["timeout_minutes"] or 0)))
    db.commit()
    db.refresh(step)
    return step


@router.delete("/epic/sub-actions/steps/{step_id}")
def delete_sub_step(step_id: int, db: Session = Depends(get_db)):
    step = db.query(EpicSubActionStep).filter(EpicSubActionStep.id == step_id).first()
    if not step:
        raise HTTPException(status_code=404, detail="Step not found")
    db.delete(step)
    db.commit()
    return {"ok": True}


# ─── Epic Sub-Action Outcomes ───

@router.get("/epic/sub-actions/{sub_id}/outcomes")
def list_sub_outcomes(sub_id: int, db: Session = Depends(get_db)):
    return db.query(EpicSubActionOutcome).filter(EpicSubActionOutcome.sub_action_id == sub_id).all()


@router.put("/epic/sub-actions/outcomes/{outcome_id}")
async def update_sub_outcome(outcome_id: int, request: Request, db: Session = Depends(get_db)):
    outcome = db.query(EpicSubActionOutcome).filter(EpicSubActionOutcome.id == outcome_id).first()
    if not outcome:
        raise HTTPException(status_code=404, detail="Outcome not found")
    b = await _json_body(request)
    if "label" in b: outcome.label = b["label"]
    if "moodlet_title" in b: outcome.moodlet_title = b["moodlet_title"]
    if "moodlet_text" in b: outcome.moodlet_text = b["moodlet_text"]
    if "image_path" in b: outcome.image_path = b["image_path"]
    db.commit()
    db.refresh(outcome)
    return outcome


# ─── Epic Action Outcomes (simple actions) ───

@router.put("/epic/actions/outcomes/{outcome_id}")
async def update_action_outcome(outcome_id: int, request: Request, db: Session = Depends(get_db)):
    outcome = db.query(EpicActionOutcome).filter(EpicActionOutcome.id == outcome_id).first()
    if not outcome:
        raise HTTPException(status_code=404, detail="Outcome not found")
    b = await _json_body(request)
    if "label" in b: outcome.label = b["label"]
    if "moodlet_title" in b: outcome.moodlet_title = b["moodlet_title"]
    if "moodlet_text" in b: outcome.moodlet_text = b["moodlet_text"]
    if "image_path" in b: outcome.image_path = b["image_path"]
    db.commit()
    db.refresh(outcome)
    return outcome


# ─── Legend fragments (легендарные драконы, phase=1) ───

@router.get("/dragons/{dragon_id}/legend")
def get_legend(dragon_id: int, db: Session = Depends(get_db)):
    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")
    fragments = (
        db.query(DragonStep)
        .filter(DragonStep.dragon_id == dragon_id, DragonStep.phase == 1)
        .order_by(DragonStep.step_number)
        .all()
    )
    return {
        "legend_image_path": dragon.legend_image_path or "",
        "legend_title": dragon.legend_title or "",
        "legend_full_text": dragon.legend_full_text or "",
        "fragments": fragments,
    }


@router.put("/dragons/{dragon_id}/legend")
async def save_legend(dragon_id: int, request: Request, db: Session = Depends(get_db)):
    dragon = db.query(Dragon).filter(Dragon.id == dragon_id).first()
    if not dragon:
        raise HTTPException(status_code=404, detail="Dragon not found")
    data = await _json_body(request)
    if "legend_image_path" in data:
        dragon.legend_image_path = data["legend_image_path"]
    if "legend_title" in data:
        dragon.legend_title = data["legend_title"]
    if "legend_full_text" in data:
        dragon.legend_full_text = data["legend_full_text"]
    fragments = data.get("fragments", [])
    submitted_ids = {f.get("id", 0) for f in fragments if f.get("id", 0) != 0}
    if submitted_ids:
        db.query(DragonStep).filter(
            DragonStep.dragon_id == dragon_id,
            DragonStep.phase == 1,
            DragonStep.id.notin_(submitted_ids),
        ).delete(synchronize_session=False)
    else:
        db.query(DragonStep).filter(
            DragonStep.dragon_id == dragon_id, DragonStep.phase == 1
        ).delete(synchronize_session=False)
    for i, f in enumerate(fragments, start=1):
        th = max(0, int(f.get("timeout_hours", 0) or 0))
        tm = max(0, min(59, int(f.get("timeout_minutes", 0) or 0)))
        cn = max(1, int(f.get("crosses_norm", 1000) or 1000))
        fid = f.get("id", 0)
        if fid == 0:
            db.add(DragonStep(
                dragon_id=dragon_id, step_number=i, phase=1,
                task_description=f.get("task_description", ""),
                magic_action=f.get("magic_action", ""),
                image_path=f.get("image_path", ""),
                keyword="вышито", timeout_hours=th, timeout_minutes=tm, crosses_norm=cn,
            ))
        else:
            frag = db.query(DragonStep).filter(DragonStep.id == fid, DragonStep.dragon_id == dragon_id).first()
            if frag:
                frag.step_number = f.get("step_number", i)
                frag.task_description = f.get("task_description", frag.task_description)
                frag.magic_action = f.get("magic_action", frag.magic_action)
                frag.image_path = f.get("image_path", frag.image_path)
                frag.timeout_hours = th
                frag.timeout_minutes = tm
                frag.crosses_norm = cn
    db.commit()
    return {"ok": True, "saved": len(fragments)}


# ─── Suspicious reports ───

@router.get("/suspicious")
def list_suspicious(status: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(SuspiciousReport)
    if status:
        q = q.filter(SuspiciousReport.status == status)
    return q.order_by(SuspiciousReport.id.desc()).limit(500).all()


@router.get("/suspicious/detailed")
def detailed_suspicious(db: Session = Depends(get_db)):
    import config
    reports = db.query(SuspiciousReport).order_by(SuspiciousReport.id.desc()).limit(500).all()
    names = _resolve_vk_names(list({r.user_id for r in reports}))
    group_id = config.VK_GROUP_ID
    items = []
    for r in reports:
        nm = names.get(r.user_id, {})
        full = " ".join(x for x in [nm.get("first_name", ""), nm.get("last_name", "")] if x).strip()
        chat_url = f"https://vk.com/gim{group_id}/convo/{r.user_id}" if group_id else f"https://vk.com/id{r.user_id}"
        items.append({
            "id": r.id,
            "user_id": r.user_id,
            "name": full or f"id{r.user_id}",
            "chat_url": chat_url,
            "message": r.raw_message or "",
            "declared_crosses": r.declared_crosses,
            "normal_crosses": r.normal_crosses,
            "mode": r.mode,
            "step_number": r.step_number,
            "created_at": r.created_at,
        })
    return {"total": db.query(SuspiciousReport).count(), "items": items}



@router.get("/suspicious/recent")
def recent_suspicious(limit: int = Query(20), db: Session = Depends(get_db)):
    reports = (
        db.query(SuspiciousReport)
        .filter(SuspiciousReport.status == "pending")
        .order_by(SuspiciousReport.id.desc())
        .limit(max(1, min(100, limit)))
        .all()
    )
    total = db.query(SuspiciousReport).filter(SuspiciousReport.status == "pending").count()
    names = _resolve_vk_names(list({r.user_id for r in reports}))
    dragon_ids = list({r.dragon_id for r in reports if r.dragon_id})
    dragon_map = {}
    if dragon_ids:
        dragon_map = {
            d.id: (d.name or d.egg_type or f"#{d.id}")
            for d in db.query(Dragon).filter(Dragon.id.in_(dragon_ids)).all()
        }
    items = []
    for r in reports:
        nm = names.get(r.user_id, {})
        full = " ".join(x for x in [nm.get("first_name", ""), nm.get("last_name", "")] if x).strip()
        items.append({
            "id": r.id,
            "user_id": r.user_id,
            "name": full or f"id{r.user_id}",
            "dragon_id": r.dragon_id,
            "dragon_name": dragon_map.get(r.dragon_id) if r.dragon_id else None,
            "step_number": r.step_number,
            "declared_crosses": r.declared_crosses,
            "normal_crosses": r.normal_crosses,
            "mode": r.mode,
            "created_at": r.created_at,
        })
    return {"total_pending": total, "items": items}


@router.delete("/suspicious/{report_id}")
def delete_suspicious(report_id: int, db: Session = Depends(get_db)):
    report = db.query(SuspiciousReport).filter(SuspiciousReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    db.delete(report)
    db.commit()
    return {"ok": True}


# ─── Piggy bank manual adjust ───

@router.post("/users/{vk_id}/balance")
async def adjust_balance(vk_id: int, request: Request, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    b = await _json_body(request)
    if "balance" in b:
        user.stitches_balance = max(0, int(b["balance"] or 0))
    elif "delta" in b:
        user.stitches_balance = max(0, (user.stitches_balance or 0) + int(b["delta"] or 0))
    else:
        raise HTTPException(status_code=400, detail="balance or delta required")
    db.commit()
    return {"ok": True, "stitches_balance": user.stitches_balance}


# ─── Магазин: цена и наборы (Robokassa) ───

@router.get("/pricing")
def get_pricing(db: Session = Depends(get_db)):
    from services.payment_service import get_base_price
    return {"base_price_rub": get_base_price(db) // 100}


@router.put("/pricing")
async def update_pricing(request: Request, db: Session = Depends(get_db)):
    from services.payment_service import set_base_price
    b = await _json_body(request)
    if "base_price_rub" not in b:
        raise HTTPException(status_code=400, detail="base_price_rub required")
    rub = max(0, int(b["base_price_rub"] or 0))
    set_base_price(db, rub * 100)
    return {"base_price_rub": rub}


@router.get("/sets")
def list_sets(db: Session = Depends(get_db)):
    return db.query(DragonSet).order_by(DragonSet.id).all()


@router.post("/sets")
async def create_set(request: Request, db: Session = Depends(get_db)):
    b = await _json_body(request)
    name = (b.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    quantity = int(b.get("quantity", 0) or 0)
    if quantity < 1:
        raise HTTPException(status_code=400, detail="quantity must be >= 1")
    s = DragonSet(
        name=name,
        quantity=quantity,
        discount_percent=max(0, int(b.get("discount_percent", 0) or 0)),
        donor_discount_percent=max(0, int(b.get("donor_discount_percent", 0) or 0)),
        is_active=bool(b.get("is_active", True)),
        created_at=datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


@router.put("/sets/{set_id}")
async def update_set(set_id: int, request: Request, db: Session = Depends(get_db)):
    s = db.query(DragonSet).filter(DragonSet.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")
    b = await _json_body(request)
    if "name" in b: s.name = b["name"]
    if "quantity" in b: s.quantity = max(1, int(b["quantity"] or 1))
    if "discount_percent" in b: s.discount_percent = max(0, int(b["discount_percent"] or 0))
    if "donor_discount_percent" in b: s.donor_discount_percent = max(0, int(b["donor_discount_percent"] or 0))
    if "is_active" in b: s.is_active = bool(b["is_active"])
    db.commit()
    db.refresh(s)
    return s


@router.delete("/sets/{set_id}")
def delete_set(set_id: int, db: Session = Depends(get_db)):
    s = db.query(DragonSet).filter(DragonSet.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")
    s.is_active = False
    db.commit()
    return {"ok": True}


@router.get("/payment-orders")
def list_payment_orders(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    status: Optional[str] = Query(None, regex="^(pending|success|fail)?$"),
    db: Session = Depends(get_db),
):
    q = db.query(PaymentOrder).outerjoin(
        DragonSet, PaymentOrder.set_id == DragonSet.id,
    ).add_columns(DragonSet.name.label("set_name"))
    if status:
        q = q.filter(PaymentOrder.status == status)
    total = q.count()
    items = q.order_by(PaymentOrder.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
    result = []
    for order, set_name in items:
        d = {
            "id": order.id,
            "vk_id": order.vk_id,
            "set_id": order.set_id,
            "set_name": set_name or "",
            "amount_rub": order.amount_rub,
            "quantity": order.quantity,
            "price_per_pin": order.price_per_pin,
            "robokassa_inv_id": order.robokassa_inv_id,
            "status": order.status,
            "dragon_ids": json.loads(order.dragon_ids or "[]"),
            "notified": order.notified,
            "created_at": order.created_at,
            "completed_at": order.completed_at,
        }
        result.append(d)
    return {"items": result, "total": total, "page": page, "per_page": per_page}


@router.post("/users/{vk_id}/custom-price")
async def set_custom_price(vk_id: int, request: Request, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.vk_id == vk_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    b = await _json_body(request)
    val = b.get("custom_price_per_dragon")
    if val is None or val == "":
        user.custom_price_per_dragon = None
    else:
        user.custom_price_per_dragon = max(0, int(val) * 100)
    db.commit()
    db.refresh(user)
    return {"custom_price_per_dragon": user.custom_price_per_dragon}


_INTRO_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "images", "intro")


def _save_intro_image(file: UploadFile) -> str:
    os.makedirs(_INTRO_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename or ".png")[1].lower() or ".png"
    from time import time
    ts = int(time())
    filename = f"intro_{ts}{ext}"
    path = os.path.join(_INTRO_DIR, filename)
    with open(path, "wb") as f:
        f.write(file.file.read())
    return f"intro/{filename}"


@router.get("/intro-chapters")
def list_intro_chapters(db: Session = Depends(get_db)):
    return db.query(IntroChapter).order_by(IntroChapter.chapter_number).all()


@router.post("/intro-chapters")
def create_intro_chapter(
    chapter_number: int = Form(...),
    text: str = Form(""),
    is_active: bool = Form(True),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    existing = db.query(IntroChapter).filter(IntroChapter.chapter_number == chapter_number).first()
    if existing:
        raise HTTPException(status_code=400, detail="Chapter with this number already exists")
    ch = IntroChapter(chapter_number=chapter_number, text=text, is_active=is_active)
    if image and image.filename:
        ch.image_path = _save_intro_image(image)
    db.add(ch)
    db.commit()
    db.refresh(ch)
    return ch


@router.put("/intro-chapters/{chapter_id}")
def update_intro_chapter(
    chapter_id: int,
    chapter_number: Optional[int] = Form(None),
    text: Optional[str] = Form(None),
    is_active: Optional[bool] = Form(None),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    ch = db.query(IntroChapter).filter(IntroChapter.id == chapter_id).first()
    if not ch:
        raise HTTPException(status_code=404, detail="Chapter not found")
    if chapter_number is not None:
        dup = db.query(IntroChapter).filter(
            IntroChapter.chapter_number == chapter_number,
            IntroChapter.id != chapter_id,
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail="Chapter with this number already exists")
        ch.chapter_number = chapter_number
    if text is not None:
        ch.text = text
    if is_active is not None:
        ch.is_active = is_active
    if image and image.filename:
        ch.image_path = _save_intro_image(image)
    db.commit()
    db.refresh(ch)
    return ch


@router.delete("/intro-chapters/{chapter_id}")
def delete_intro_chapter(chapter_id: int, db: Session = Depends(get_db)):
    ch = db.query(IntroChapter).filter(IntroChapter.id == chapter_id).first()
    if not ch:
        raise HTTPException(status_code=404, detail="Chapter not found")
    db.delete(ch)
    db.commit()
    return {"ok": True}
