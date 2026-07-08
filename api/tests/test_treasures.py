from models import Dragon, DragonStep, User, UserDragon, Treasure, UserTreasure, UserLegendProgress


def _rare_dragon(db, name="Rare", rarity=2):
    d = Dragon(name=name, rarity=rarity, steps_count=1, is_active=True)
    db.add(d)
    db.commit()
    return d


def test_upsert_treasure_and_list(client, db):
    d = _rare_dragon(db)

    resp = client.post(f"/api/admin/dragons/{d.id}/treasure", data={"name": "Gem", "description": "Shiny"})
    assert resp.status_code == 200
    body = resp.json()
    assert body["name"] == "Gem"
    assert body["dragon_id"] == d.id

    resp2 = client.post(f"/api/admin/dragons/{d.id}/treasure", data={"name": "Gem2", "description": "New"})
    assert resp2.status_code == 200
    assert db.query(Treasure).filter(Treasure.dragon_id == d.id).count() == 1

    lst = client.get("/api/admin/treasures")
    assert lst.status_code == 200
    assert len(lst.json()) == 1
    assert lst.json()[0]["name"] == "Gem2"


def test_treasure_rarity_validation(client, db):
    d = _rare_dragon(db, rarity=1)
    resp = client.post(f"/api/admin/dragons/{d.id}/treasure", data={"name": "X", "description": ""})
    assert resp.status_code == 400


def test_dragon_detail_includes_treasure(client, db):
    d = _rare_dragon(db)
    client.post(f"/api/admin/dragons/{d.id}/treasure", data={"name": "Gem", "description": "d"})
    resp = client.get(f"/api/admin/dragons/{d.id}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["treasure"] is not None
    assert body["treasure"]["name"] == "Gem"


def test_update_and_delete_treasure(client, db):
    d = _rare_dragon(db)
    created = client.post(f"/api/admin/dragons/{d.id}/treasure", data={"name": "Gem", "description": "d"}).json()
    tid = created["id"]

    upd = client.put(f"/api/admin/treasures/{tid}", data={"name": "Edited"})
    assert upd.status_code == 200
    assert upd.json()["name"] == "Edited"

    dele = client.delete(f"/api/admin/treasures/{tid}")
    assert dele.status_code == 200
    assert db.query(Treasure).count() == 0


def test_collection_treasures_endpoint(client, db):
    d1 = _rare_dragon(db, name="R1")
    d2 = _rare_dragon(db, name="R2")
    t1 = Treasure(name="Owned", description="mine", image_path="dragons/a.png", dragon_id=d1.id, is_active=True)
    t2 = Treasure(name="Missing", description="hidden", image_path="dragons/b.png", dragon_id=d2.id, is_active=True)
    db.add_all([t1, t2])
    db.add(User(vk_id=42))
    db.flush()
    db.add(UserTreasure(user_id=42, treasure_id=t1.id))
    db.commit()

    resp = client.get("/api/collection/42/treasures")
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 2
    assert len(body["collected"]) == 1
    assert body["collected"][0]["name"] == "Owned"
    assert len(body["uncollected"]) == 1
    assert "silhouette" in body["uncollected"][0]


def test_collection_legends_endpoint(client, db):
    d = Dragon(name="Legendary", rarity=3, steps_count=1, is_active=True, legend_image_path="dragons/cover.png")
    db.add(d)
    db.flush()
    db.add_all([
        DragonStep(dragon_id=d.id, step_number=1, phase=1),
        DragonStep(dragon_id=d.id, step_number=2, phase=1),
    ])
    db.add(User(vk_id=99))
    db.flush()
    db.add(UserLegendProgress(user_id=99, dragon_id=d.id, fragment_number=1, completed=True))
    db.commit()

    resp = client.get("/api/collection/99/legends")
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["dragon_id"] == d.id
    assert body[0]["fragments_total"] == 2
    assert body[0]["fragments_opened"] == 1


def test_reset_dragon_removes_user_treasures(client, db):
    d = _rare_dragon(db)
    t = Treasure(name="Gem", dragon_id=d.id, is_active=True)
    db.add(t)
    db.add(User(vk_id=10, current_dragon_id=d.id))
    db.flush()
    db.add(UserDragon(user_id=10, dragon_id=d.id, completed_at=""))
    db.add(UserTreasure(user_id=10, treasure_id=t.id))
    db.commit()

    resp = client.post("/api/admin/users/10/reset-dragon", json={"dragon_id": d.id})
    assert resp.status_code == 200
    assert db.query(UserTreasure).filter(UserTreasure.user_id == 10).count() == 0


def test_delete_user_dragon_removes_user_treasures(client, db):
    d = _rare_dragon(db)
    t = Treasure(name="Gem", dragon_id=d.id, is_active=True)
    db.add(t)
    db.add(User(vk_id=11))
    db.flush()
    db.add(UserDragon(user_id=11, dragon_id=d.id, completed_at="2026-01-01T00:00:00"))
    db.add(UserTreasure(user_id=11, treasure_id=t.id))
    db.commit()

    resp = client.delete("/api/admin/users/11/dragons/" + str(d.id))
    assert resp.status_code == 200
    assert db.query(UserTreasure).filter(UserTreasure.user_id == 11).count() == 0


def test_export_dragons_xlsx(client, db):
    import io
    from openpyxl import load_workbook

    d = Dragon(name="Exportable", rarity=2, egg_type="ice", steps_count=2, is_active=True)
    db.add(d)
    db.flush()
    db.add_all([
        DragonStep(dragon_id=d.id, step_number=1, phase=0, magic_action="A1", task_description="T1", crosses_norm=1000),
        DragonStep(dragon_id=d.id, step_number=2, phase=0, magic_action="A2", task_description="T2", crosses_norm=1500),
    ])
    db.add(Treasure(name="Gem", description="Shiny", dragon_id=d.id, is_active=True))
    db.commit()

    resp = client.get("/api/admin/dragons/export")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert "Драконы" in wb.sheetnames
    assert "Шаги" in wb.sheetnames
    d_rows = list(wb["Драконы"].iter_rows(values_only=True))
    assert d_rows[0][1] == "Имя"
    assert any(r[1] == "Exportable" and r[9] == "Gem" for r in d_rows[1:])
    s_rows = list(wb["Шаги"].iter_rows(values_only=True))
    assert len([r for r in s_rows[1:] if r[1] == "Exportable"]) == 2


def test_export_dragons_excludes_epic(client, db):
    import io
    from openpyxl import load_workbook

    db.add(Dragon(name="Regular", rarity=1, steps_count=0, is_active=True, is_epic=False))
    db.add(Dragon(name="EpicOne", rarity=3, steps_count=0, is_active=True, is_epic=True))
    db.commit()

    resp = client.get("/api/admin/dragons/export")
    wb = load_workbook(io.BytesIO(resp.content))
    names = [r[1] for r in wb["Драконы"].iter_rows(min_row=2, values_only=True)]
    assert "Regular" in names
    assert "EpicOne" not in names


def test_user_detail_includes_collected_treasures(client, db):
    d1 = _rare_dragon(db, name="R1")
    d2 = _rare_dragon(db, name="R2")
    t1 = Treasure(name="Gem", description="Shiny", dragon_id=d1.id, is_active=True)
    t2 = Treasure(name="Coin", description="Gold", dragon_id=d2.id, is_active=True)
    db.add_all([t1, t2])
    db.add(User(vk_id=77))
    db.flush()
    db.add(UserTreasure(user_id=77, treasure_id=t1.id))
    db.commit()

    resp = client.get("/api/admin/users/77")
    assert resp.status_code == 200
    body = resp.json()
    assert "treasures_collected" in body
    assert len(body["treasures_collected"]) == 1
    assert body["treasures_collected"][0]["name"] == "Gem"


def test_available_dragons_excludes_rare_with_treasure(client, db):
    rare_with = _rare_dragon(db, name="WithTreasure")
    rare_without = _rare_dragon(db, name="WithoutTreasure")
    common = _rare_dragon(db, name="Common", rarity=1)
    legendary = _rare_dragon(db, name="Legendary", rarity=3)
    db.add(Treasure(name="Gem", dragon_id=rare_with.id, is_active=True))
    db.commit()

    resp = client.get("/api/admin/treasures/available-dragons")
    assert resp.status_code == 200
    names = [d["name"] for d in resp.json()]
    assert "WithoutTreasure" in names
    assert "WithTreasure" not in names
    assert "Common" not in names
    assert "Legendary" not in names


def test_available_dragons_empty_when_all_taken(client, db):
    d = _rare_dragon(db, name="OnlyOne")
    db.add(Treasure(name="Gem", dragon_id=d.id, is_active=True))
    db.commit()

    resp = client.get("/api/admin/treasures/available-dragons")
    assert resp.status_code == 200
    assert resp.json() == []
